import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
import io
import time

class ExcelGenerator:
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # 移除默认工作表
    
    def add_data_sheet(self, data: pd.DataFrame, sheet_name: str = "数据") -> None:
        """
        添加数据工作表
        :param data: 要写入Excel的数据
        :param sheet_name: 工作表名称
        """
        ws = self.wb.create_sheet(title=sheet_name)
        
        # 写入数据
        for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                # 设置表头样式
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def add_summary_sheet(self, summary_data: pd.DataFrame, sheet_name: str = "统计摘要") -> None:
        """
        添加统计摘要工作表
        :param summary_data: 统计数据
        :param sheet_name: 工作表名称
        """
        ws = self.wb.create_sheet(title=sheet_name)
        
        for r_idx, row in enumerate(dataframe_to_rows(summary_data, index=True, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                # 设置表头和索引样式
                if r_idx == 1 or c_idx == 1:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
    
    def add_chart(self, data: pd.DataFrame, chart_type: str, title: str, sheet_name: str, 
                 x_column: str = None, y_columns: list = None, chart_position: str = "A10") -> None:
        """
        添加图表到指定工作表
        :param data: 数据
        :param chart_type: 图表类型 (bar, line, pie)
        :param title: 图表标题
        :param sheet_name: 目标工作表名称
        :param x_column: X轴列名
        :param y_columns: Y轴列名列表
        :param chart_position: 图表在工作表中的位置
        """
        ws = self.wb[sheet_name]
        
        # 如果没有指定列，使用默认列
        if x_column is None:
            x_column = data.columns[0]
        if y_columns is None:
            y_columns = data.columns[1:3] if len(data.columns) > 1 else [data.columns[0]]
        
        # 获取数据范围
        data_start_row = 1
        data_end_row = len(data) + 1
        x_col_idx = data.columns.get_loc(x_column) + 1
        
        # 创建图表
        if chart_type == "bar":
            chart = BarChart()
            chart.type = "col"
        elif chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        else:
            raise ValueError("不支持的图表类型，请使用 'bar', 'line' 或 'pie'")
        
        chart.title = title
        chart.style = 10
        
        # 设置数据
        for y_col in y_columns:
            y_col_idx = data.columns.get_loc(y_col) + 1
            data_ref = Reference(ws, min_col=y_col_idx, min_row=data_start_row, max_row=data_end_row)
            cat_ref = Reference(ws, min_col=x_col_idx, min_row=data_start_row + 1, max_row=data_end_row)
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cat_ref)
        
        # 将图表添加到工作表
        ws.add_chart(chart, chart_position)
    
    def add_matplotlib_chart(self, data: pd.DataFrame, chart_type: str, title: str, sheet_name: str, 
                            x_column: str = None, y_columns: list = None, chart_position: str = "A10",
                            figsize: tuple = (10, 6)) -> None:
        """
        使用matplotlib生成图表并插入到Excel中
        :param data: 数据
        :param chart_type: 图表类型 (bar, line, pie, scatter, histogram)
        :param title: 图表标题
        :param sheet_name: 目标工作表名称
        :param x_column: X轴列名
        :param y_columns: Y轴列名列表
        :param chart_position: 图表在工作表中的位置
        :param figsize: 图表大小
        """
        ws = self.wb[sheet_name]
        
        # 如果没有指定列，使用默认列
        if x_column is None:
            x_column = data.columns[0]
        if y_columns is None:
            y_columns = data.columns[1:3] if len(data.columns) > 1 else [data.columns[0]]
        
        # 创建图表
        plt.figure(figsize=figsize)
        plt.title(title)
        
        if chart_type == "bar":
            data.plot(x=x_column, y=y_columns, kind="bar", ax=plt.gca())
        elif chart_type == "line":
            data.plot(x=x_column, y=y_columns, kind="line", ax=plt.gca())
        elif chart_type == "pie":
            if len(y_columns) == 1:
                data.set_index(x_column)[y_columns[0]].plot(kind="pie", ax=plt.gca(), autopct='%1.1f%%')
        elif chart_type == "scatter":
            if len(y_columns) > 0:
                for y_col in y_columns:
                    plt.scatter(data[x_column], data[y_col], label=y_col)
                plt.legend()
        elif chart_type == "histogram":
            if len(y_columns) > 0:
                for y_col in y_columns:
                    plt.hist(data[y_col].dropna(), alpha=0.5, label=y_col)
                plt.legend()
        else:
            raise ValueError("不支持的图表类型")
        
        plt.tight_layout()
        
        # 保存图表到内存
        img_stream = io.BytesIO()
        plt.savefig(img_stream, format='png', dpi=150)
        img_stream.seek(0)
        plt.close()
        
        # 插入图表到Excel
        from openpyxl.drawing.image import Image
        img = Image(img_stream)
        ws.add_image(img, chart_position)
    
    def save(self, file_path: str) -> None:
        """
        保存Excel文件
        :param file_path: 保存路径
        """
        self.wb.save(file_path)
    
    def generate_excel(self, data: pd.DataFrame, output_path: str = "report.xlsx") -> str:
        """
        生成完整的Excel报告
        :param data: 输入数据
        :param output_path: 输出路径
        :return: 生成的文件路径
        """
        start_time = time.time()
        
        # 添加数据工作表
        self.add_data_sheet(data, "数据")
        
        # 添加统计摘要
        summary_data = data.describe()
        self.add_summary_sheet(summary_data, "统计摘要")
        
        # 添加图表
        numeric_cols = data.select_dtypes(include=['number']).columns.tolist()
        if len(numeric_cols) > 0:
            # 柱状图
            self.add_matplotlib_chart(data, "bar", "柱状图分析", "数据", 
                                     x_column=data.columns[0], y_columns=numeric_cols[:3], 
                                     chart_position="E1")
            
            # 折线图
            self.add_matplotlib_chart(data, "line", "折线图分析", "数据", 
                                     x_column=data.columns[0], y_columns=numeric_cols[:3], 
                                     chart_position="E15")
        
        # 保存文件
        self.save(output_path)
        
        end_time = time.time()
        print(f"Excel生成耗时: {end_time - start_time:.2f}秒")
        
        return output_path
